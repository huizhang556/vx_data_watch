from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

ACCOUNT_HEADERS = ["时间", "播放", "推荐", "喜欢", "评论", "分享", "关注"]


class ImportValidationError(ValueError):
    pass


@dataclass
class AccountMetricRow:
    metric_date: date
    plays: int
    recommendations: int | None
    likes: int | None
    comments: int | None
    shares: int | None
    follows: int | None
    raw: dict[str, str]

    def normalized(self) -> dict[str, Any]:
        result = asdict(self)
        result["metric_date"] = self.metric_date.isoformat()
        return result


def file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _parse_nonnegative(value: Any, field: str, allow_empty: bool = True) -> int | None:
    if value is None or str(value).strip() == "":
        if allow_empty:
            return None
        raise ImportValidationError(f"{field}不能为空")
    cleaned = str(value).strip().replace(",", "")
    if not re.fullmatch(r"\d+", cleaned):
        raise ImportValidationError(f"{field}必须是非负整数，收到：{value}")
    return int(cleaned)


def parse_account_csv(content: bytes) -> list[AccountMetricRow]:
    if b"\x00" in content:
        raise ImportValidationError("文件不是有效的文本 CSV")
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportValidationError("CSV 必须使用 UTF-8 编码") from exc
    rows = list(csv.reader(io.StringIO(text)))
    header_index = next(
        (
            i
            for i, row in enumerate(rows)
            if set(ACCOUNT_HEADERS).issubset({cell.strip() for cell in row})
        ),
        None,
    )
    if header_index is None:
        raise ImportValidationError("未找到预期表头：时间、播放、推荐、喜欢、评论、分享、关注")

    headers = [cell.strip() for cell in rows[header_index]]
    parsed: list[AccountMetricRow] = []
    for line_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(cell.strip() for cell in values):
            continue
        raw = {
            header: (values[i].strip() if i < len(values) else "")
            for i, header in enumerate(headers)
        }
        try:
            metric_date = datetime.strptime(raw["时间"], "%Y/%m/%d").date()
            parsed.append(
                AccountMetricRow(
                    metric_date=metric_date,
                    plays=_parse_nonnegative(raw["播放"], "播放", allow_empty=False) or 0,
                    recommendations=_parse_nonnegative(raw.get("推荐"), "推荐"),
                    likes=_parse_nonnegative(raw.get("喜欢"), "喜欢"),
                    comments=_parse_nonnegative(raw.get("评论"), "评论"),
                    shares=_parse_nonnegative(raw.get("分享"), "分享"),
                    follows=_parse_nonnegative(raw.get("关注"), "关注"),
                    raw=raw,
                )
            )
        except (ValueError, ImportValidationError) as exc:
            raise ImportValidationError(f"第 {line_number} 行：{exc}") from exc
    if not parsed:
        raise ImportValidationError("CSV 中没有数据行")
    if len({row.metric_date for row in parsed}) != len(parsed):
        raise ImportValidationError("CSV 中存在重复日期")
    return parsed


VIDEO_HEADER_ALIASES = {
    "metric_date": ["数据日期", "日期", "metric_date"],
    "title": ["视频标题", "标题", "title"],
    "published_at": ["发布时间", "published_at"],
    "plays": ["当日播放量", "昨日新增播放", "播放", "plays"],
    "identity_key": ["视频标识", "视频ID", "identity_key"],
    "cumulative_plays": ["累计播放量", "cumulative_plays"],
    "likes": ["喜欢", "likes"],
    "comments": ["评论", "comments"],
    "shares": ["分享", "shares"],
}


def _find_alias(headers: list[str], canonical: str) -> str | None:
    return next((alias for alias in VIDEO_HEADER_ALIASES[canonical] if alias in headers), None)


def parse_video_sheet(content: bytes, filename: str) -> list[dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        raw_rows = list(csv.DictReader(io.StringIO(text)))
    elif filename.lower().endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            raise ImportValidationError("工作表为空")
        headers = [str(cell or "").strip() for cell in values[0]]
        raw_rows = [
            dict(zip(headers, row, strict=False))
            for row in values[1:]
            if any(cell is not None for cell in row)
        ]
    else:
        raise ImportValidationError("逐视频文件仅支持 CSV 或 XLSX")
    if not raw_rows:
        raise ImportValidationError("文件中没有数据行")
    headers = [str(key).strip() for key in raw_rows[0]]
    mapping = {key: _find_alias(headers, key) for key in VIDEO_HEADER_ALIASES}
    for required in ("metric_date", "title", "plays"):
        if not mapping[required]:
            raise ImportValidationError(f"缺少字段：{VIDEO_HEADER_ALIASES[required][0]}")

    parsed: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_rows, start=2):
        try:
            raw_date = str(raw[mapping["metric_date"]]).strip()  # type: ignore[index]
            metric_date = datetime.fromisoformat(raw_date.replace("/", "-")).date()
            title = str(raw[mapping["title"]]).strip()  # type: ignore[index]
            if not title:
                raise ImportValidationError("视频标题不能为空")
            published_at = None
            if mapping["published_at"] and raw.get(mapping["published_at"]):
                published_at = datetime.fromisoformat(
                    str(raw[mapping["published_at"]]).strip().replace("/", "-")
                )
            parsed.append(
                {
                    "metric_date": metric_date,
                    "title": title,
                    "published_at": published_at,
                    "plays": _parse_nonnegative(raw[mapping["plays"]], "当日播放量", False),  # type: ignore[index]
                    "identity_key": str(raw.get(mapping["identity_key"], "") or "").strip() or None,
                    "cumulative_plays": _parse_nonnegative(
                        raw.get(mapping["cumulative_plays"]), "累计播放量"
                    )
                    if mapping["cumulative_plays"]
                    else None,
                    "likes": _parse_nonnegative(raw.get(mapping["likes"]), "喜欢")
                    if mapping["likes"]
                    else None,
                    "comments": _parse_nonnegative(raw.get(mapping["comments"]), "评论")
                    if mapping["comments"]
                    else None,
                    "shares": _parse_nonnegative(raw.get(mapping["shares"]), "分享")
                    if mapping["shares"]
                    else None,
                    "raw": {str(k): str(v or "") for k, v in raw.items()},
                }
            )
        except (ValueError, ImportValidationError) as exc:
            raise ImportValidationError(f"第 {index} 行：{exc}") from exc
    return parsed


def normalize_title(title: str) -> str:
    return re.sub(r"[\s…\.，,。!！?？:：;；'\"“”‘’]+", "", title).lower()


def video_identity(title: str, published_at: datetime | None, explicit: str | None = None) -> str:
    if explicit:
        return explicit[:180]
    seed = f"{normalize_title(title)}|{published_at.isoformat() if published_at else ''}"
    return "auto:" + hashlib.sha256(seed.encode("utf-8")).hexdigest()[:40]
